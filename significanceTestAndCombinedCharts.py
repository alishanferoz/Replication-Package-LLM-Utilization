import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
from matplotlib.patches import Patch
import re
import os
import textwrap
from scipy.stats import fisher_exact
from openpyxl.styles import PatternFill
from openpyxl import load_workbook

# ---------------------------------------------------------
# 1. CONFIGURATION & MAPPING
# ---------------------------------------------------------

INPUT_FILE = 'results-survey797953.xlsx'
OUTPUT_SIG_FILE = "Significance test analysis_Outpu.xlsx"

FREQ_SCALE_USAGE = ["Never", "Rarely", "Sometimes", "Often", "Always"]
FREQ_SCALE_BENEFIT = ["No benefit", "Low benefit", "Moderate benefit", "High benefit", "Very high benefit"]

# Consistent Color Palettes
COLORS_USAGE = ['#F0F6FB', '#BDD7E7', '#6BAED6', '#2171B5', '#08306B']
COLORS_BENEFIT = ['#F2F0F7', '#CBC9E2', '#9E9AC8', '#6A51A3', '#3F007D']

# Mapping for significance tests
TEXT_MAPPING_SIG = {
    "never": "Never", "rarely": "Rarely", "sometimes": "Sometimes", "sometime": "Sometimes",
    "often": "Often", "always": "Always", "no benefit": "No benefit", "low benefit": "Low benefit",
    "moderate benefit": "Moderate benefit", "high benefit": "High benefit", "very high benefit": "Very high benefit"
}

SHEET_MAPPING = {
    "Requirements": "In which of the following requirement-gathering tasks",
    "Software Design": "In which of the following software designing tasks",
    "Development": "In which of the following development tasks",
    "Testing": "In which of the following software testing tasks",
    "Methods": "How often do you use the following methods",
    "Integration Benefit": "How much benefit do you get from the integration of LLM"
}

DISTRIBUTE_BENEFIT_STEM = "How much do you benefit from the outputs of LLMs for the following tasks"


# ---------------------------------------------------------
# 2. UTILITY FUNCTIONS
# ---------------------------------------------------------

def get_text_color(hex_color):
    try:
        rgb = mcolors.hex2color(hex_color)
        luminance = 0.299 * rgb[0] + 0.587 * rgb[1] + 0.114 * rgb[2]
        return 'black' if luminance > 0.5 else 'white'
    except:
        return 'black'


def clean_text_value(val):
    if pd.isna(val) or val == "": return np.nan
    val = str(val).strip()
    val = re.sub(r'\s*\(.*?\)', '', val)
    v_l = val.lower()
    for k, v in TEXT_MAPPING_SIG.items():
        if v_l == k: return v
    return val


def extract_subtask(col_name):
    if '[?]' in col_name:
        content = col_name.split('[?]')[0].split('[')[-1]
    elif '[' in col_name and ']' in col_name:
        content = re.findall(r'\[(.*?)\]', col_name)[-1]
    elif '?' in col_name:
        parts = col_name.split('?')
        content = parts[-1] if len(parts) > 1 and parts[-1].strip() else col_name
    else:
        content = col_name
    return content.split('(')[0].split('[')[0].strip()


def format_task_label(text):
    words = text.split()
    if not words: return ""
    res = [words[0]]
    for i in range(1, len(words)):
        word, prev = words[i], words[i - 1]
        delim = " " if (len(word) == 3 or len(prev) <= 2) else "\n"
        res.append(delim + word)
    return "".join(res)


# ---------------------------------------------------------
# 3. SIGNIFICANCE ENGINE
# ---------------------------------------------------------

def run_significance_analysis(df_full):
    print("Running Fisher's Exact Tests...")
    writer = pd.ExcelWriter(OUTPUT_SIG_FILE, engine='openpyxl')

    sig_data_cache = {}

    for sheet_name, stem in SHEET_MAPPING.items():
        cols = [c for c in df_full.columns if stem.lower() in c.lower()]
        if not cols: continue

        sample = df_full[cols[0]].dropna()
        is_benefit = not sample.empty and "benefit" in str(sample.iloc[0]).lower()
        scale = FREQ_SCALE_BENEFIT if is_benefit else FREQ_SCALE_USAGE

        rows = []
        for c in cols:
            task = extract_subtask(c)
            counts = df_full[c].apply(clean_text_value).value_counts()
            row = {"Task": task}
            for s in scale: row[s] = counts.get(s, 0)
            rows.append(row)

        df_table = pd.DataFrame(rows).set_index("Task")

        if sheet_name in ["Requirements", "Software Design", "Development", "Testing"]:
            ben_cols = [c for c in df_full.columns if DISTRIBUTE_BENEFIT_STEM.lower() in c.lower()]
            ben_rows = []
            for bc in ben_cols:
                if extract_subtask(bc) in df_table.index:
                    counts = df_full[bc].apply(clean_text_value).value_counts()
                    brow = {"Task": extract_subtask(bc)}
                    for s in FREQ_SCALE_BENEFIT: brow[s] = counts.get(s, 0)
                    ben_rows.append(brow)
            df_ben = pd.DataFrame(ben_rows).set_index("Task")
            sig_data_cache[sheet_name] = {'usage': df_table, 'benefit': df_ben}
        else:
            sig_data_cache[sheet_name] = {'usage': df_table}

    for s_name, data in sig_data_cache.items():
        ws = writer.book.create_sheet(s_name)
        curr_r = 1
        for key in ['usage', 'benefit']:
            if key in data and not data[key].empty:
                df = data[key]
                ws.cell(curr_r, 1, f"{s_name} {key.capitalize()} Table")
                df.to_excel(writer, sheet_name=s_name, startrow=curr_r)

                grand_total = df.sum().sum()
                p_matrix = pd.DataFrame(index=df.index, columns=df.columns)

                for task in df.index:
                    for col in df.columns:
                        obs = df.loc[task, col]
                        row_sum, col_sum = df.loc[task].sum(), df[col].sum()
                        table = np.array([[obs, row_sum - obs], [col_sum - obs, grand_total - row_sum - col_sum + obs]])
                        _, p = fisher_exact(table)
                        p_matrix.loc[task, col] = p

                curr_r += len(df) + 3
                ws.cell(curr_r, 1, f"P-Values for {key}")
                p_matrix.to_excel(writer, sheet_name=s_name, startrow=curr_r)
                curr_r += len(p_matrix) + 3

    writer.close()
    return sig_data_cache


# ---------------------------------------------------------
# 4. CHARTING ENGINE WITH SIGNIFICANCE MARKERS
# ---------------------------------------------------------

def get_sig_label(task, category, sig_data, data_type):
    """
    Checks if a task/category pair is significant.
    - SH: Observed > Expected
    - SL: Observed < Expected
    """
    if not sig_data or data_type not in sig_data: return ""
    df = sig_data[data_type]
    if task not in df.index or category not in df.columns: return ""

    obs = df.loc[task, category]
    row_sum = df.loc[task].sum()
    col_sum = df[category].sum()
    grand_total = df.sum().sum()
    if grand_total == 0: return ""

    # Expected value calculation
    exp = (row_sum * col_sum) / grand_total

    table = np.array([[obs, row_sum - obs], [col_sum - obs, grand_total - row_sum - col_sum + obs]])
    _, p = fisher_exact(table)

    if p < 0.05:
        return "(SH)" if obs > exp else "(SL)"
    return ""


# def plot_combined_stacked(df, matched_data, sig_info, title, filename):
#     n_tasks = len(matched_data)
#     step, bar_h, gap = 0.45, 0.18, 0.02
#     y_idx = np.arange(n_tasks) * step
#     fig, ax = plt.subplots(figsize=(14, n_tasks * 1.8 + 2))
#
#     def draw_group(y_pos, col_key, scale, colors, data_type):
#         lefts = np.zeros(n_tasks)
#         for i, cat in enumerate(scale):
#             widths, counts = [], []
#             for item in matched_data:
#                 c_counts = df[item[col_key]].value_counts()
#                 c = c_counts.get(cat, 0)
#                 total = sum([c_counts.get(k, 0) for k in scale])
#                 widths.append((c / total * 100) if total > 0 else 0)
#                 counts.append(c)
#
#             bars = ax.barh(y_pos, widths, left=lefts, height=bar_h, color=colors[i], edgecolor='white', linewidth=0.5)
#             lefts += np.array(widths)
#
#             for j, rect in enumerate(bars):
#                 w = rect.get_width()
#                 if w > 5:
#                     task_name = matched_data[j]['task_orig']
#                     marker = get_sig_label(task_name, cat, sig_info, data_type)
#                     text_color = get_text_color(colors[i])
#
#                     # Construct a single label with all parts stacked
#                     # This ensures they remain perfectly centered as a group
#                     full_label = f"{w:.0f}%\n({counts[j]})"
#                     if marker:
#                         full_label += f"\n{marker}"
#
#                     # Determine text color for significance (red if it's SH/SL, else standard text color)
#                     # To keep it simple and clean, we'll draw the whole block
#                     # If we need the marker to be red specifically, we handle it via split text or just text color
#
#                     ax.text(rect.get_x() + w / 2, rect.get_y() + rect.get_height() / 2,
#                             full_label,
#                             ha='center', va='center', color=text_color,
#                             fontsize=16, fontweight='bold')
#
#     draw_group(y_idx - (bar_h / 2 + gap / 2), 'u_col', FREQ_SCALE_USAGE, COLORS_USAGE, 'usage')
#     draw_group(y_idx + (bar_h / 2 + gap / 2), 'b_col', FREQ_SCALE_BENEFIT, COLORS_BENEFIT, 'benefit')
#
#     ax.set_yticks(y_idx)
#     ax.set_yticklabels([format_task_label(m['task_clean']) for m in matched_data], fontsize=16, fontweight='bold')
#     ax.invert_yaxis()
#     ax.margins(x=0)
#     ax.tick_params(axis='y', pad=55)
#
#     for y in y_idx:
#         ax.text(-1, y - 0.1, "Usage", ha='right', va='center', fontsize=14, fontstyle='italic', color='#444')
#         ax.text(-1, y + 0.1, "Benefit", ha='right', va='center', fontsize=14, fontstyle='italic', color='#444')
#
#     leg1 = ax.legend(handles=[Patch(facecolor=c, label=l) for c, l in zip(COLORS_USAGE, FREQ_SCALE_USAGE)],
#                      title="Usage Frequency", loc='upper left', bbox_to_anchor=(1.02, 1.0), fontsize=14)
#     ax.add_artist(leg1)
#     ax.legend(handles=[Patch(facecolor=c, label=l) for c, l in zip(COLORS_BENEFIT, FREQ_SCALE_BENEFIT)],
#               title="Benefit Level", loc='upper left', bbox_to_anchor=(1.02, 0.45), fontsize=14)
#
#     plt.tight_layout()
#     plt.subplots_adjust(right=0.75)
#     plt.savefig(filename, dpi=300)
#     plt.close()
def plot_combined_stacked(df, matched_data, sig_info, title, filename):
    n_tasks = len(matched_data)
    step, bar_h, gap = 0.45, 0.18, 0.02
    y_idx = np.arange(n_tasks) * step
    fig, ax = plt.subplots(figsize=(14, n_tasks * 1.8 + 2))

    def draw_group(y_pos, col_key, scale, colors, data_type):
        lefts = np.zeros(n_tasks)
        for i, cat in enumerate(scale):
            widths, counts = [], []
            for item in matched_data:
                c_counts = df[item[col_key]].value_counts()
                c = c_counts.get(cat, 0)
                total = sum([c_counts.get(k, 0) for k in scale])
                widths.append((c / total * 100) if total > 0 else 0)
                counts.append(c)

            bars = ax.barh(y_pos, widths, left=lefts, height=bar_h, color=colors[i], edgecolor='white', linewidth=0.5)
            lefts += np.array(widths)

            for j, rect in enumerate(bars):
                w = rect.get_width()
                if w > 5:
                    task_name = matched_data[j]['task_orig']
                    marker = get_sig_label(task_name, cat, sig_info, data_type)
                    text_color = get_text_color(colors[i])

                    # Exact center coordinates
                    x_mid = rect.get_x() + w / 2
                    y_mid = rect.get_y() + rect.get_height() / 2

                    # 1. Draw the primary data (Percentage and Count)
                    # We add a trailing newline if a marker exists to keep the vertical spacing identical
                    main_label = f"{w:.0f}%\n({counts[j]})"
                    if marker:
                        main_label += "\n "  # Add a blank line for the marker space

                    ax.text(x_mid, y_mid, main_label,
                            ha='center', va='center', color=text_color,
                            fontsize=16, fontweight='bold')

                    # 2. Draw ONLY the marker in Red
                    # We use leading newlines to push the marker to the third line position
                    if marker:
                        marker_label = f"\n\n{marker}"
                        ax.text(x_mid, y_mid, marker_label,
                                ha='center', va='center', color='red',
                                fontsize=16, fontweight='bold')

    draw_group(y_idx - (bar_h / 2 + gap / 2), 'u_col', FREQ_SCALE_USAGE, COLORS_USAGE, 'usage')
    draw_group(y_idx + (bar_h / 2 + gap / 2), 'b_col', FREQ_SCALE_BENEFIT, COLORS_BENEFIT, 'benefit')

    ax.set_yticks(y_idx)
    ax.set_yticklabels([format_task_label(m['task_clean']) for m in matched_data], fontsize=16, fontweight='bold')
    ax.invert_yaxis()
    ax.margins(x=0)
    ax.tick_params(axis='y', pad=55)

    for y in y_idx:
        ax.text(-1, y - 0.1, "Usage", ha='right', va='center', fontsize=14, fontstyle='italic', color='#444')
        ax.text(-1, y + 0.1, "Benefit", ha='right', va='center', fontsize=14, fontstyle='italic', color='#444')

    leg1 = ax.legend(handles=[Patch(facecolor=c, label=l) for c, l in zip(COLORS_USAGE, FREQ_SCALE_USAGE)],
                     title="Usage Frequency", loc='upper left', bbox_to_anchor=(1.02, 1.0), fontsize=14)
    ax.add_artist(leg1)
    ax.legend(handles=[Patch(facecolor=c, label=l) for c, l in zip(COLORS_BENEFIT, FREQ_SCALE_BENEFIT)],
              title="Benefit Level", loc='upper left', bbox_to_anchor=(1.02, 0.45), fontsize=14)

    plt.tight_layout()
    plt.subplots_adjust(right=0.75)
    plt.savefig(filename, dpi=300)
    plt.close()


# ---------------------------------------------------------
# 5. MAIN EXECUTION
# ---------------------------------------------------------

def main():
    df = pd.read_excel(INPUT_FILE)
    for col in df.columns: df[col] = df[col].apply(clean_text_value)

    # Step 1: Run Significance
    sig_cache = run_significance_analysis(df)

    # Step 2: Prepare Charts
    ben_tasks = [c for c in df.columns if DISTRIBUTE_BENEFIT_STEM in c]

    tasks_configs = [
        ("Requirements", "In which of the following requirement-gathering tasks", "chart_E_req_combined.png"),
        ("Software Design", "In which of the following software designing tasks", "chart_F_design_combined.png"),
        ("Development", "In which of the following development tasks", "chart_F_dev_combined.png"),
        ("Testing", "In which of the following software testing tasks", "chart_F_test_combined.png")
    ]

    print("Generating Charts with Significance Markers...")
    for s_key, stem, fname in tasks_configs:
        u_cols = [c for c in df.columns if stem in c]
        matched = []
        for uc in u_cols:
            u_task = extract_subtask(uc)
            for bc in ben_tasks:
                if extract_subtask(bc) == u_task:
                    counts = df[uc].value_counts()
                    w = (counts.get("Often", 0) + counts.get("Always", 0)) / counts.sum() if counts.sum() > 0 else 0
                    matched.append({'task_orig': u_task, 'task_clean': u_task, 'u_col': uc, 'b_col': bc, 'weight': w})
                    break

        if matched:
            matched = sorted(matched, key=lambda x: x['weight'], reverse=True)
            plot_combined_stacked(df, matched, sig_cache.get(s_key), "", fname)

    # Access Method Chart
    acc_cols = [c for c in df.columns if "How often do you use the following methods" in c]
    acc_ben_cols = [c for c in df.columns if "benefit do you get from the integration" in c]
    acc_matched = []
    for ac in acc_cols:
        a_task = extract_subtask(ac)
        for abc in acc_ben_cols:
            if extract_subtask(abc) == a_task:
                counts = df[ac].value_counts()
                w = (counts.get("Often", 0) + counts.get("Always", 0)) / counts.sum() if counts.sum() > 0 else 0
                acc_matched.append({'task_orig': a_task, 'task_clean': a_task, 'u_col': ac, 'b_col': abc, 'weight': w})

    if acc_matched:
        acc_matched = sorted(acc_matched, key=lambda x: x['weight'], reverse=True)
        plot_combined_stacked(df, acc_matched, sig_cache.get("Methods"), "", "chart_F_access_combined.png")

    print("Analysis and Visualization Complete.")


if __name__ == "__main__":
    main()