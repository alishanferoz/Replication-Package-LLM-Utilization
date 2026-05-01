import pandas as pd
import os
import re
import numpy as np
from scipy.stats import fisher_exact
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# Configuration
input_file = "results-survey797953.xlsx"
output_file = "Significance test analysis_Output.xlsx"

# 1. Define Categories and their specific Question Stems
sheet_mapping = {
    "Requirements": "In which of the following requirement-gathering tasks",
    "Software Design": "In which of the following software designing tasks",
    "Development": "In which of the following development tasks",
    "Testing": "In which of the following software testing tasks",
    "Methods": "How often do you use the following methods",
    "Integration Benefit": "How much benefit do you get from the integration of LLM"
}

DISTRIBUTE_BENEFIT_QUESTION = "How much do you benefit from the outputs of LLMs for the following tasks"

# Mapping dictionary
text_mapping = {
    # Frequency
    "never": "Never",
    "rarely": "Rarely",
    "sometimes": "Sometime",
    "sometime": "Sometime",
    "often": "Often",
    "always": "Always",
    # Benefit
    "no benefit": "No benefit",
    "low benefit": "Low benefit",
    "moderate benefit": "Moderate benefit",
    "high benefit": "High benefit",
    "very high benefit": "Very High Benefit"
}

# Define Sort Orders
orders = {
    "Frequency": ["Never", "Rarely", "Sometime", "Often", "Always"],
    "Benefit": ["No benefit", "Low benefit", "Moderate benefit", "High benefit", "Very High Benefit"]
}


def clean_response(text):
    if pd.isna(text):
        return None
    text_lower = str(text).lower()
    sorted_keys = sorted(text_mapping.keys(), key=len, reverse=True)
    for key in sorted_keys:
        if key in text_lower:
            return text_mapping[key]
    return text


def determine_scale_type(df_columns):
    if df_columns.empty: return "Frequency"
    first_col = df_columns.iloc[:, 0].dropna()
    if first_col.empty: return "Frequency"
    sample_text = str(first_col.iloc[0]).lower()
    if "benefit" in sample_text: return "Benefit"
    return "Frequency"


def extract_subtask(col_name):
    # Priority 1: Handle [?] marker
    if '[?]' in col_name:
        pre_marker = col_name.split('[?]')[0]
        if '[' in pre_marker:
            content = pre_marker.split('[')[-1]
        else:
            content = pre_marker

    # Priority 2: Standard [...] format
    elif '[' in col_name and ']' in col_name:
        matches = re.findall(r'\[(.*?)\]', col_name)
        if matches:
            content = matches[-1]
        else:
            content = col_name

    # Priority 3: Question? Option format
    elif '?' in col_name:
        parts = col_name.split('?')
        if len(parts) > 1 and parts[-1].strip():
            content = parts[-1]
        else:
            content = col_name
    else:
        content = col_name

    # Clean (e.g., descriptions in parens)
    content = content.split('(')[0]
    content = content.split('[')[0]  # Extra safety for nested brackets

    return content.strip()


def extract_data_for_stem(df, question_stem):
    matching_cols = [c for c in df.columns if question_stem.lower() in c.lower()]

    if not matching_cols:
        print(f"Warning: No columns found for '{question_stem}'")
        return pd.DataFrame()

    scale_type = determine_scale_type(df[matching_cols])
    scale_order = orders[scale_type]

    rows = []

    for col in matching_cols:
        task_name = extract_subtask(col)
        counts = df[col].apply(clean_response).value_counts()

        row_data = {"Task": task_name}
        for item in scale_order:
            row_data[item] = counts.get(item, 0)

        rows.append(row_data)

    df_result = pd.DataFrame(rows)
    if not df_result.empty:
        df_result.set_index("Task", inplace=True)
        df_result = df_result[scale_order]

    return df_result


def get_color_gradient(value, min_val, max_val):
    """
    Returns the hex color for a value based on a Light Green -> Dark Green gradient.
    Returns None if value is 0 (to skip coloring).
    """
    if value == 0:
        return None

    # Colors: Light Green (E2EFDA) to Dark Green (228B22)
    # RGB values
    start_rgb = (226, 239, 218)  # E2EFDA
    end_rgb = (34, 139, 34)  # 228B22

    if max_val <= min_val:
        # Avoid division by zero if all non-zero values are the same
        ratio = 1.0
    else:
        ratio = (value - min_val) / (max_val - min_val)

    r = int(start_rgb[0] + (end_rgb[0] - start_rgb[0]) * ratio)
    g = int(start_rgb[1] + (end_rgb[1] - start_rgb[1]) * ratio)
    b = int(start_rgb[2] + (end_rgb[2] - start_rgb[2]) * ratio)

    return f"{r:02X}{g:02X}{b:02X}"


def perform_fisher_exact_analysis(writer, worksheet, df, sheet_name, start_row):
    """
    Generates contingency tables for EACH CELL, calculates Fisher's Exact Test,
    and writes Summary Matrices + Detailed Tables.
    """
    grand_total = df.sum().sum()

    # Initialize empty dataframes for stats and p-values with same shape as df
    df_stats = pd.DataFrame(index=df.index, columns=df.columns)
    df_pvalues = pd.DataFrame(index=df.index, columns=df.columns)

    current_row = start_row

    # Header for Detailed Tables section
    worksheet.cell(row=current_row, column=1, value="Detailed Contingency Tables (Per Cell)")
    current_row += 2

    # Iterate through every cell (Row=Task, Col=ScaleOption)
    for task in df.index:
        for col_name in df.columns:

            current = df.loc[task, col_name]
            row_sum = df.loc[task].sum()
            col_sum = df[col_name].sum()

            right = row_sum - current
            bottom = col_sum - current
            bottom_right = grand_total - row_sum - col_sum + current

            # Construct 2x2 Contingency Table
            contingency_table = np.array([
                [current, right],
                [bottom, bottom_right]
            ])

            # Calculate Fisher's Exact Test
            if contingency_table.sum() > 0:
                odds_ratio, p = fisher_exact(contingency_table, alternative='two-sided')
            else:
                odds_ratio, p = 0.0, 1.0

            # Store results in the summary matrices (Odds Ratio and P-value)
            df_stats.loc[task, col_name] = round(odds_ratio, 2)
            df_pvalues.loc[task, col_name] = round(p, 4)

            # Write the individual 2x2 table (Detailed Output)
            worksheet.cell(row=current_row, column=1, value=f"Contingency table for {task} x {col_name}")

            df_small = pd.DataFrame(
                contingency_table,
                index=[task, f"Not {task}"],
                columns=[col_name, f"Not {col_name}"]
            )
            df_small.to_excel(writer, sheet_name=sheet_name, startrow=current_row)

            # Move down (Header + 2 rows + space)
            current_row += 5

    # Now write the Summary Matrices at the bottom
    current_row += 2

    # 1. Fisher's Exact Odds Ratio Matrix
    worksheet.cell(row=current_row, column=1, value="Fisher's Exact Stat Matrix")
    current_row += 1
    df_stats.to_excel(writer, sheet_name=sheet_name, startrow=current_row)
    current_row += len(df_stats) + 3

    # 2. Fisher's Exact Odds Ratio Matrix with whole number values (NEW)
    worksheet.cell(row=current_row, column=1, value="Fisher's Exact Stat Matrix with whole number values")
    current_row += 1

    # Round to nearest whole number and convert to integer
    # FIX: Handle inf/nan values properly and avoid downcasting warning by converting to numeric first
    df_stats_numeric = df_stats.apply(pd.to_numeric, errors='coerce')
    df_stats_whole = df_stats_numeric.replace([np.inf, -np.inf], 0).fillna(0).round(0).astype(int)
    df_stats_whole.to_excel(writer, sheet_name=sheet_name, startrow=current_row)

    # --- Color Formatting Logic (Manual Application to Exclude 0) ---
    all_values = df_stats_whole.values.flatten()
    # Filter for non-zero values to determine min/max for the gradient
    non_zero_values = all_values[all_values > 0]

    if len(non_zero_values) > 0:
        min_val = non_zero_values.min()
        max_val = non_zero_values.max()

        data_start_row = current_row + 2

        # Iterate over the DataFrame dimensions to color specific cells
        for r_idx in range(len(df_stats_whole)):
            for c_idx in range(len(df_stats_whole.columns)):
                val = df_stats_whole.iloc[r_idx, c_idx]

                hex_color = get_color_gradient(val, min_val, max_val)

                if hex_color:
                    # Column starts at 2 because Column 1 is the Index (Task Names)
                    cell = worksheet.cell(row=data_start_row + r_idx, column=2 + c_idx)
                    cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')

    current_row += len(df_stats_whole) + 3

    # 3. Fisher's Exact P-Values Matrix
    worksheet.cell(row=current_row, column=1, value="Fisher's Exact P-Values Matrix")
    current_row += 1
    df_pvalues.to_excel(writer, sheet_name=sheet_name, startrow=current_row)

    return current_row + len(df_pvalues) + 4


# --- EXECUTION ---

print(f"Reading '{input_file}'...")
if input_file.endswith('.csv'):
    df_full = pd.read_csv(input_file)
else:
    df_full = pd.read_excel(input_file)

# Data Store
sheets_data = {name: {'usage': pd.DataFrame(), 'benefit': pd.DataFrame()} for name in sheet_mapping.keys()}
task_to_sheet_map = {}

# 1. Process Usage
print("Processing Usage Questions...")
for sheet_name, stem in sheet_mapping.items():
    df_extracted = extract_data_for_stem(df_full, stem)
    if not df_extracted.empty:
        sheets_data[sheet_name]['usage'] = df_extracted
        for task in df_extracted.index:
            task_to_sheet_map[task] = sheet_name

# 2. Process Benefit Distribution
print("Processing Benefit Distribution...")
benefit_cols = [c for c in df_full.columns if DISTRIBUTE_BENEFIT_QUESTION.lower() in c.lower()]
scale_order_benefit = orders["Benefit"]
rows_by_sheet = {name: [] for name in sheet_mapping.keys()}
rows_by_sheet["Misc"] = []

for col in benefit_cols:
    task_name = extract_subtask(col)
    target_sheet = task_to_sheet_map.get(task_name, "Misc")
    counts = df_full[col].apply(clean_response).value_counts()

    row_data = {"Task": task_name}
    for item in scale_order_benefit:
        row_data[item] = counts.get(item, 0)

    rows_by_sheet[target_sheet].append(row_data)

for sheet, rows in rows_by_sheet.items():
    if rows:
        if sheet not in sheets_data:
            sheets_data[sheet] = {'usage': pd.DataFrame(), 'benefit': pd.DataFrame()}
        df_ben = pd.DataFrame(rows)
        df_ben.set_index("Task", inplace=True)
        df_ben = df_ben[scale_order_benefit]
        sheets_data[sheet]['benefit'] = df_ben

# 3. Write Output
writer = pd.ExcelWriter(output_file, engine='openpyxl')

for sheet_name, data in sheets_data.items():
    usage_df = data['usage']
    benefit_df = data['benefit']

    if usage_df.empty and benefit_df.empty:
        continue

    start_row = 0

    # Write Usage
    if not usage_df.empty:
        writer.book.create_sheet(sheet_name)
        worksheet = writer.sheets[sheet_name]
        worksheet.cell(row=start_row + 1, column=1, value=f"{sheet_name} Usage table")
        usage_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row + 1)

        start_row = start_row + len(usage_df) + 4
        # Perform Per-Cell Analysis
        start_row = perform_fisher_exact_analysis(writer, worksheet, usage_df, sheet_name, start_row)
        start_row += 2
    else:
        if sheet_name not in writer.sheets:
            writer.book.create_sheet(sheet_name)
        worksheet = writer.sheets[sheet_name]

    # Write Benefit
    if not benefit_df.empty:
        worksheet.cell(row=start_row + 1, column=1, value=f"{sheet_name} Benefit table")
        benefit_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row + 1)

        start_row = start_row + len(benefit_df) + 4
        # Perform Per-Cell Analysis
        start_row = perform_fisher_exact_analysis(writer, worksheet, benefit_df, sheet_name, start_row)

    # Formatting
    worksheet.column_dimensions['A'].width = 50
    for col in ['B', 'C', 'D', 'E', 'F', 'G']:
        worksheet.column_dimensions[col].width = 18

if 'Sheet' in writer.book.sheetnames and len(writer.book.sheetnames) > 1:
    del writer.book['Sheet']

writer.close()
print(f"Analysis complete with Per-Cell Fisher's Exact statistics. File saved to: {output_file}")